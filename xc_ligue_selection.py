#!/usr/bin/env python3
"""Simple utility to load two Excel files into pandas DataFrames.

This provides:
- load_excel_file(path): loads a single Excel file with openpyxl engine
- load_both(...): convenience to load the two files used by the project
- a CLI `main()` which prints shape() and head() of each DataFrame

If the files are missing the script exits with a non-zero code and prints
an explanatory message to stderr.
"""
from pathlib import Path
import argparse
import sys
import difflib
import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime


def load_excel_file(path, engine='openpyxl', sheet_name=0):
	"""Load a single Excel file into a pandas DataFrame.

	Raises FileNotFoundError if the file does not exist and RuntimeError on
	other read errors.
	"""
	p = Path(path)
	if not p.exists():
		raise FileNotFoundError(f"File not found: {p.resolve()}")
	try:
		return pd.read_excel(p, engine=engine, sheet_name=sheet_name)
	except Exception as e:
		raise RuntimeError(f"Failed to read Excel file {p}: {e}") from e


def load_both(base_dir='.', signup_file='liga_signups.xlsx', rangliste_file='rangliste.xlsx', sheet_name=0):
	"""Load both expected Excel files and return (df_signups, df_rangliste)."""
	base = Path(base_dir)
	s_path = base / signup_file
	r_path = base / rangliste_file
	df_s = load_excel_file(s_path, sheet_name=sheet_name)
	df_r = load_excel_file(r_path, sheet_name=sheet_name)
	return df_s, df_r


def extract_columns(df, columns=('Vorname', 'Nachname', 'sum')):
	"""Return a DataFrame with only the requested summary columns from rangliste.

	Parameters
	- df: pandas.DataFrame loaded from the rangliste Excel file
	- columns: iterable of column names to extract (default: 'Vorname','Nachname','sum')

	Returns a new DataFrame with only those columns in the given order.

	Raises ValueError if any expected column is missing from the DataFrame.
	"""
	# Normalize columns tuple -> list to preserve order
	cols = list(columns)
	missing = [c for c in cols if c not in df.columns]
	if missing:
		raise ValueError(f"Missing expected columns in rangliste DataFrame: {missing}")
	return df.loc[:, cols].copy()


def drop_rows_where_xc_none(df, xc_col='XC'):
	"""Return a copy of df excluding rows where `xc_col` is None/NaN.

	This does NOT remove empty strings ('' ) — it only filters out true
	missing values (pd.isna). If you also want to treat empty strings as
	missing, call df[xc_col].replace({'': pd.NA}) before passing it here.

	Raises ValueError if the column does not exist.
	"""
	if xc_col not in df.columns:
		raise ValueError(f"Column not found in DataFrame: {xc_col}")
	# keep rows where XC is not NA (None/NaN)
	return df[df[xc_col].notna()].copy()


def _norm(s):
	"""Normalize a value to a lowercase stripped string for matching."""
	return (str(s) if pd.notna(s) else '').strip().lower()


def _build_lookups(signups_df, signup_first_col, signup_last_col, columns_to_add):
	"""Return (_signups, right, right_lookup, full_lookup, right_cols).

	- _signups: copy of signups_df with '__key'
	- right: DataFrame indexed by __key with the requested right-side columns
	- right_lookup: dict mapping __key -> {col: val}
	- full_lookup: dict mapping 'firstname lastname' -> (__key, {col: val})
	- right_cols: list of columns actually present to bring across
	"""
	_signups = signups_df.copy()
	_signups['__key'] = _signups[signup_first_col].apply(_norm) + '|' + _signups[signup_last_col].apply(_norm)

	# which columns are available on the right side
	right_cols = [c for c in columns_to_add if c in _signups.columns]

	if right_cols:
		right = _signups.set_index('__key')[right_cols].groupby(level=0).first()
	else:
		right = pd.DataFrame(columns=columns_to_add)

	right_lookup = right.to_dict(orient='index') if not right.empty else {}

	# full lookup uses normalized "firstname lastname"
	full_lookup = {}
	for _, su in _signups.iterrows():
		f_s = _norm(su.get(signup_first_col))
		l_s = _norm(su.get(signup_last_col))
		full_key = f_s + ' ' + l_s
		if full_key not in full_lookup:
			full_lookup[full_key] = (su.get('__key'), {c: su.get(c) for c in right_cols})

	return _signups, right, right_lookup, full_lookup, right_cols


def _try_fill_matches(joined, full_lookup, right_lookup, _signups, right_cols,
					  rang_name_cols, signup_first_col, signup_last_col,
					  fuzzy_interactive=False, fuzzy_cutoff=0.85):
	"""Attempt swapped, full-name, and fuzzy suggestions to fill missing right-side cols.

	Modifies `joined` in place.
	"""
	if not right_cols:
		return

	candidates = list(full_lookup.keys()) if full_lookup else []

	for idx, prow in joined.iterrows():
		missing_cols = [c for c in right_cols if pd.isna(prow.get(c))]
		if not missing_cols:
			continue

		f = prow.get(rang_name_cols[0])
		l = prow.get(rang_name_cols[1])

		# swapped key: lastname|firstname
		swapped_key = _norm(l) + '|' + _norm(f)
		if swapped_key in right_lookup:
			vals = right_lookup[swapped_key]
			for c in right_cols:
				joined.at[idx, c] = vals.get(c)
			joined.at[idx, '__key'] = swapped_key
			continue

		# full-name exact
		left_full = _norm(f) + ' ' + _norm(l)
		if left_full in full_lookup:
			sign_key, vals = full_lookup[left_full]
			for c in right_cols:
				joined.at[idx, c] = vals.get(c)
			if sign_key:
				joined.at[idx, '__key'] = sign_key
			continue

		# fuzzy suggestion (hint-only or interactive)
		if candidates:
			matches = difflib.get_close_matches(left_full, candidates, n=1, cutoff=fuzzy_cutoff)
			if matches:
				matched = matches[0]
				sign_key, vals = full_lookup.get(matched)
				# try to fetch original capitalization for display
				orig_first = None
				orig_last = None
				row_match = _signups[_signups['__key'] == sign_key]
				if not row_match.empty:
					orig_first = row_match.iloc[0].get(signup_first_col)
					orig_last = row_match.iloc[0].get(signup_last_col)
				display_name = f"{orig_first} {orig_last}" if (orig_first or orig_last) else matched
				print(f"Suggestion: close signup name for '{prow.get(rang_name_cols[0])} {prow.get(rang_name_cols[1])}' -> '{display_name}'")
				accepted = False
				if fuzzy_interactive:
					try:
						resp = input("Accept suggestion and use these signup values? [y/n]: ").strip()
					except (EOFError, KeyboardInterrupt):
						resp = ''
					# accept on yes/y (case-insensitive), decline on no/n; anything else -> decline
					resp_norm = resp.lower()
					if resp_norm in ('y', 'yes'):
						accepted = True
					elif resp_norm in ('n', 'no'):
						accepted = False
					else:
						accepted = False
				if accepted:
					for c in right_cols:
						joined.at[idx, c] = vals.get(c)
					if sign_key:
						joined.at[idx, '__key'] = sign_key


def _ensure_columns(joined, columns_to_add):
	for col in columns_to_add:
		if col not in joined.columns:
			joined[col] = None


def _format_birthdate(joined):
	if 'Birthdate' in joined.columns:
		_bd = pd.to_datetime(joined['Birthdate'], errors='coerce')
		joined['Birthdate'] = _bd.dt.strftime('%d.%m.%Y')
		joined['Birthdate'] = joined['Birthdate'].where(joined['Birthdate'].notna(), None)


def _append_unmatched(joined, _signups, rang_name_cols, signup_first_col, signup_last_col, columns_to_add):
	_signup_keys = set(_signups['__key'].dropna())
	_matched_keys = set(joined['__key'].dropna())
	_unmatched_keys = _signup_keys - _matched_keys
	if not _unmatched_keys:
		return joined.drop(columns='__key')

	_unmatched = _signups[_signups['__key'].isin(_unmatched_keys)].copy()
	out_cols = [c for c in joined.columns if c != '__key']
	rows = []
	for _, u in _unmatched.iterrows():
		row = {}
		for c in out_cols:
			if c == rang_name_cols[0]:
				row[c] = u.get(signup_first_col)
			elif c == rang_name_cols[1]:
				row[c] = u.get(signup_last_col)
			elif c == 'sum':
				row[c] = 0
			elif c in columns_to_add:
				row[c] = u.get(c) if c in u.index else None
			else:
				row[c] = None
		rows.append(row)
	df_unmatched = pd.DataFrame(rows, columns=out_cols)
	if 'Birthdate' in df_unmatched.columns:
		_bd2 = pd.to_datetime(df_unmatched['Birthdate'], errors='coerce')
		df_unmatched['Birthdate'] = _bd2.dt.strftime('%d.%m.%Y')
		df_unmatched['Birthdate'] = df_unmatched['Birthdate'].where(df_unmatched['Birthdate'].notna(), None)
	df_unmatched = df_unmatched.sort_values(by=rang_name_cols[1], key=lambda s: s.str.lower()).reset_index(drop=True)
	return pd.concat([joined.drop(columns='__key'), df_unmatched], ignore_index=True)



def combine_rangliste_with_signups(rang_df, signups_df,
								   rang_name_cols=('Vorname', 'Nachname'),
								   signup_first_col='Firstname', signup_last_col='Lastname',
								   columns_to_add=['Sex', 'XC', 'Birthdate'],
								   fuzzy_interactive=False, fuzzy_cutoff=0.85):
	"""Combine rangliste and signups preserving rangliste order.

	For each row in `rang_df` (expected to have columns named by
	`rang_name_cols`), look up a matching row in `signups_df` where
	`signup_first_col` == Vorname and `signup_last_col` == Nachname (case-
	insensitive, stripped). If a match is found the returned DataFrame will
	contain the requested `sex_col` and `xc_col` values (if present in
	signups_df). If no match is found those fields will be None.

	This function preserves the order of `rang_df`.

	Returns a new DataFrame (a copy) with the added columns 
	"""
	# validate inputs
	if rang_name_cols[0] not in rang_df.columns or rang_name_cols[1] not in rang_df.columns:
		raise ValueError(f"rang_df must contain columns: {rang_name_cols}")
	if signup_first_col not in signups_df.columns or signup_last_col not in signups_df.columns:
		raise ValueError(f"signups_df must contain columns: {signup_first_col}, {signup_last_col}")

	# build lookups and helper structures
	_signups, right, right_lookup, full_lookup, right_cols = _build_lookups(
		signups_df, signup_first_col, signup_last_col, columns_to_add
	)

	# prepare left frame and join on key to preserve order
	left = rang_df.copy()
	left['__key'] = left[rang_name_cols[0]].apply(_norm) + '|' + left[rang_name_cols[1]].apply(_norm)
	joined = left.join(right, on='__key')

	# attempt to fill missing right-side values using swapped/full/fuzzy strategies
	_try_fill_matches(joined, full_lookup, right_lookup, _signups, right_cols,
					  rang_name_cols, signup_first_col, signup_last_col,
					  fuzzy_interactive=fuzzy_interactive, fuzzy_cutoff=fuzzy_cutoff)

	# ensure requested columns exist on the joined frame
	_ensure_columns(joined, columns_to_add)

	# normalize Birthdate format to DD.MM.YYYY if present
	_format_birthdate(joined)

	# Append signup rows that were not matched in rangliste (helper handles drop of __key)
	joined = _append_unmatched(joined, _signups, rang_name_cols, signup_first_col, signup_last_col, columns_to_add)

	return joined


def run_selection(args):
	df_signups, df_rangliste = load_both(
		base_dir=args.base_dir,
		signup_file=args.signup_file,
		rangliste_file=args.rangliste_file,
		sheet_name=args.sheet_name,
	)
	summary_rangliste = extract_columns(df_rangliste, columns=('Vorname', 'Nachname', 'sum'))
	summary_signups = drop_rows_where_xc_none(extract_columns(df_signups, columns=('Lastname', 'Firstname', 'Sex', 'XC', 'Birthdate')), "XC")

	combined = combine_rangliste_with_signups(
		summary_rangliste,
		summary_signups,
		rang_name_cols=('Vorname', 'Nachname'),
		signup_first_col='Firstname',
		signup_last_col='Lastname',
		columns_to_add=['Sex', 'XC', 'Birthdate'],
		fuzzy_interactive=getattr(args, 'interactive', True),
	)

	# save combined to a new Excel file
	output_path = Path(args.base_dir) / 'combined_selection.xlsx'
	_save_with_coloring(output_path, combined, sex_col='Sex')
	print(f"Combined selection saved to: {output_path.resolve()}")


def _save_with_coloring(path, df, sex_col='Sex', birthdate_col='Birthdate'):
	"""Save DataFrame to Excel and color rows:

	- red for pilots who are U26 in the next year (born in or after cutoff year)
	- green for pilots whose sex is 'f' or 'w' (case-insensitive)

	Red takes precedence over green when both apply.
	"""
	p = Path(path)
	# fills
	green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
	red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

	# compute U26 mask based on birth year >= (next_year - 26)
	next_year = datetime.now().year + 1
	cutoff_year = next_year - 26

	# safely parse birthdates (day-first) and compute mask
	if birthdate_col in df.columns:
		bd = pd.to_datetime(df[birthdate_col], dayfirst=True, errors='coerce')
		birth_year = bd.dt.year
		is_u26 = birth_year.ge(cutoff_year)
	else:
		is_u26 = pd.Series([False] * len(df))

	# female/woman mask
	if sex_col in df.columns:
		is_female = df[sex_col].astype(str).fillna('').str.strip().str.lower().isin(('f', 'w'))
	else:
		is_female = pd.Series([False] * len(df))

	# write DataFrame to Excel
	with pd.ExcelWriter(p, engine='openpyxl') as writer:
		df.to_excel(writer, index=False, sheet_name='Sheet1')
		writer.book.save(p)

	# reopen workbook and apply fills by row using masks computed above
	from openpyxl import load_workbook
	wb = load_workbook(p)
	ws = wb.active

	# map column names to 1-based column indices based on df.columns order
	col_index = {col: i + 1 for i, col in enumerate(df.columns)}

	# iterate DataFrame rows in order and apply fills
	for i in range(len(df)):
		row_idx = i + 2  # Excel data rows start at 2
		try:
			if bool(is_u26.iat[i]):
				# color whole row red
				for col_idx in range(1, ws.max_column + 1):
					ws.cell(row=row_idx, column=col_idx).fill = red_fill
				continue
			if bool(is_female.iat[i]):
				for col_idx in range(1, ws.max_column + 1):
					ws.cell(row=row_idx, column=col_idx).fill = green_fill
		except IndexError:
			# mismatch lengths or empty df; skip
			continue

	wb.save(p)


def main(argv=None):
	parser = argparse.ArgumentParser(description='Load two Excel files and print brief summaries.')
	parser.add_argument('--base-dir', default='.', help='Directory containing the Excel files')
	parser.add_argument('--signup-file', default='liga_signups.xlsx', help='Signup Excel filename')
	parser.add_argument('--rangliste-file', default='rangliste.xlsx', help='Rangliste Excel filename')
	parser.add_argument('--sheet-name', default=0, help='Sheet name or index to read (passed to pandas)')
	args = parser.parse_args(argv)

	try:
		run_selection(args)
	except FileNotFoundError as e:
		print(f"Error: {e}", file=sys.stderr)
		return 2
	except RuntimeError as e:
		print(f"Error reading Excel files: {e}", file=sys.stderr)
		return 3

	return 0


if __name__ == '__main__':
	raise SystemExit(main())

